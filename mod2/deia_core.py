# -*- coding: utf-8 -*-
"""
Nucleo di calcolo della Matrice di trasparenza DEIA.

Asse X - maturita' DEIA praticata  <- export Qualtrics (risposte 1-4)
Asse Y - maturita' DEIA comunicata <- content analysis dei report (foglio matrice_y)

La metodologia (mappatura domanda->tema, pesi PMI/GRANDI, regola MAT, allocazione
70/30, soglia e scala dei livelli) e' incorporata qui sotto come costante, estratta
da Matrice_trasparenza_DEIA_simulazione_v2_final.xlsx: in questo modo il calcolo
funziona con i soli due file caricati dall'utente.

Usato sia da app.py (Streamlit) sia da build_matrice.py (generatore Excel).
"""

import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SOGLIA = 2.5  # alta maturita' > 2,50; bassa <= 2,50
SI = "Sì"

# codice tema -> nome tema di reporting (ordine = ordine di presentazione)
CODICI = {
    "F": "Foundation",
    "HR": "HR",
    "E": "Employment",
    "EE": "Employees Engagement",
    "ET": "Education & Training",
    "ND": "Non-discrimination",
    "M": "Monitoring",
    "SC": "Supply Chain Diversity",
    "C": "Communities",
    "EU": "End-users",
}

# (cod. domanda, framework, sottogruppo, cod. standard, cod.1, cod.2, peso PMI, peso GRANDI)
DOMANDE = (
    ("CQ1", "CULTURA", "CULTURA E COMPORTAMENTI", "MAT", "", "", 1, 1),
    ("CQ2", "CULTURA", "CULTURA E COMPORTAMENTI", "EE", "EE", "", 3, 2.5),
    ("CQ3", "CULTURA", "CULTURA E COMPORTAMENTI", "EE/ND", "EE", "ND", 3, 2),
    ("CQ4", "CULTURA", "BENESSERE E SICUREZZA PSICOLOGICA", "ND", "ND", "", 1, 1),
    ("CQ5", "CULTURA", "BENESSERE E SICUREZZA PSICOLOGICA", "MAT", "", "", 2.5, 1.5),
    ("PQ1", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "EE", "EE", "", 1.5, 1),
    ("PQ2", "PROCESSI", "PERSONE E PROCESSI", "ET1/ET2", "ET", "", 2.5, 1.5),
    ("PQ3", "PROCESSI", "PERSONE E PROCESSI", "E2", "E", "", 1, 1),
    ("PQ4", "PROCESSI", "PERSONE E PROCESSI", "E3", "E", "", 3, 2),
    ("PQ5", "PROCESSI", "PERSONE E PROCESSI", "E1/E2", "E", "", 3, 3),
    ("PQ6", "PROCESSI", "PERSONE E PROCESSI", "HR1/HR2", "HR", "", 2, 1),
    ("PQ7", "PROCESSI", "PERSONE E PROCESSI", "ET1/ET4", "ET", "", 2, 1.5),
    ("CQ6", "CULTURA", "FORMAZIONE E CONSAPEVOLEZZA", "ET2/ET3", "ET", "", 3, 1.5),
    ("PQ8", "PROCESSI", "PERSONE E PROCESSI", "HR2/E1", "HR", "E", 2.5, 1.5),
    ("PQ9", "PROCESSI", "PERSONE E PROCESSI", "HR2/E1", "HR", "E", 2.5, 1.5),
    ("PQ10", "PROCESSI", "PERSONE E PROCESSI", "M/HR1", "M", "HR", 3, 1.5),
    ("PQ11", "PROCESSI", "PERSONE E PROCESSI", "E2/E4", "E", "", 1.5, 1),
    ("PQ12", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "E1/ND", "E", "ND", 3, 2),
    ("SQ1", "STRATEGIA", "STRATEGIA E PIANIFICAZIONE", "F", "F", "", 1.5, 1),
    ("PQ13", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "E1", "E", "", 1.5, 1),
    ("PQ14", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "E1", "E", "", 1.5, 1),
    ("PQ15", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "E1", "E", "", 1.5, 1),
    ("PQ16", "PROCESSI", "PROCESSI OPERATIVI E ACCESSIBILITÀ", "ND", "ND", "", 3, 2),
    ("SQ2", "STRATEGIA", "GOVERNANCE E LEADERSHIP", "F", "F", "", 2, 1),
    ("SQ3", "STRATEGIA", "GOVERNANCE E LEADERSHIP", "ET2/F", "ET", "F", 3, 1.5),
    ("SQ4", "STRATEGIA", "GOVERNANCE E LEADERSHIP", "MAT", "", "", 1.5, 1),
    ("SQ5", "STRATEGIA", "GOVERNANCE E LEADERSHIP", "M/F", "M", "F", 3, 1.5),
    ("SQ6", "STRATEGIA", "GOVERNANCE E LEADERSHIP", "F", "F", "", 2.5, 1.5),
    ("MQ1", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "M", "M", "", 3, 1.5),
    ("MQ2", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "EU", "EU", "", 3, 1),
    ("MQ3", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "EU/ND", "EU", "ND", 3, 1.5),
    ("MQ4", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "EU", "EU", "", 3, 2.5),
    ("MQ5", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "M/EU", "M", "EU", 3, 2),
    ("MQ6", "MERCATO", "BRAND, COMUNICAZIONE E RAPPRESENTAZIONE", "M", "M", "", 3, 2),
    ("MQ7", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU", "EU", "", 3, 1.5),
    ("MQ8", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "M", "M", "", 3, 2),
    ("MQ9", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU/M", "EU", "M", 3, 2.5),
    ("CQ7", "CULTURA", "FORMAZIONE E CONSAPEVOLEZZA", "ET/EU", "ET", "EU", 1.5, 1),
    ("MQ10", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU", "EU", "", 1.5, 1),
    ("MQ11", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU/EE", "EU", "EE", 2.5, 1.5),
    ("MQ12", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU", "EU", "", 1.5, 1),
    ("MQ13", "MERCATO", "CLIENTI, UTENTI E ACCESSIBILITÀ", "EU", "EU", "", 1.5, 1),
    ("PQ17", "PROCESSI", "INNOVAZIONE, RICERCA E SVILUPPO", "E1/EU", "E", "EU", 3, 2),
    ("PQ18", "PROCESSI", "INNOVAZIONE, RICERCA E SVILUPPO", "HR1", "HR", "", 3, 2),
    ("PQ19", "PROCESSI", "INNOVAZIONE, RICERCA E SVILUPPO", "ND", "ND", "", "", ""),
    ("MQ14", "MERCATO", "IMPATTO SOCIALE E POSIZIONAMENTO", "M", "M", "", 2.5, 1.5),
    ("MQ15", "MERCATO", "IMPATTO SOCIALE E POSIZIONAMENTO", "C", "C", "", 2.5, 1.5),
    ("MQ16", "MERCATO", "IMPATTO SOCIALE E POSIZIONAMENTO", "C", "C", "", 2.5, 2),
    ("MQ17", "MERCATO", "IMPATTO SOCIALE E POSIZIONAMENTO", "SC", "SC", "", 2, 1),
    ("SQ7", "STRATEGIA", "MISURAZIONE E IMPATTO", "EE/M", "EE", "M", 1.5, 1),
    ("SQ8", "STRATEGIA", "MISURAZIONE E IMPATTO", "M", "M", "", 2.5, 1.5),
    ("SQ9", "STRATEGIA", "STRATEGIA E PIANIFICAZIONE", "F", "F", "", 1.5, 1),
    ("SQ10", "STRATEGIA", "STRATEGIA E PIANIFICAZIONE", "F", "F", "", 1.5, 1),
    ("SQ11", "STRATEGIA", "STRATEGIA E PIANIFICAZIONE", "F/M", "F", "M", 1.5, 1),
    ("SQ12", "STRATEGIA", "STRATEGIA E PIANIFICAZIONE", "F/M", "F", "M", 3, 2),
)

CAMPI = ("cod", "framework", "sottogruppo", "standard", "cod1", "cod2",
         "peso_pmi", "peso_gr")
MAPPATURA = [dict(zip(CAMPI, d)) for d in DOMANDE]

# intestazioni attese nell'export Qualtrics (riga 1) per i metadati anagrafici
META_QUALTRICS = {"A1": "azienda", "A2": "piva", "A3": "dipendenti", "A4": "settore"}


class ErroreInput(Exception):
    """Il file caricato non ha la struttura attesa."""


def norm(s):
    """Normalizza un testo per il confronto: accenti, spazi multipli, maiuscole."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def classifica_dimensione(numero_dipendenti):
    """PMI se meno di 250 dipendenti (definizione UE), altrimenti Grande impresa."""
    numeri = [int(n) for n in re.findall(r"\d+", str(numero_dipendenti or ""))]
    return "Grande impresa" if numeri and max(numeri) >= 250 else "PMI"


def _foglio(source, nomi_attesi=()):
    """Apre il workbook e restituisce il foglio atteso, o il primo disponibile."""
    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:
        raise ErroreInput(f"file non leggibile come .xlsx ({exc})") from exc
    for nome in nomi_attesi:
        for foglio in wb.sheetnames:
            if norm(foglio) == norm(nome):
                return wb[foglio]
    return wb[wb.sheetnames[0]]


# ------------------------------------------------------------ asse X: survey
def elenca_rispondenti(source):
    """Righe compilate dell'export Qualtrics: [(numero_riga, etichetta), ...]."""
    ws = _foglio(source, ["Sheet0"])
    col_azienda = next((c for c in range(1, ws.max_column + 1)
                        if str(ws.cell(1, c).value or "").strip() == "A1"), None)
    out = []
    for r in range(3, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        nome = ws.cell(r, col_azienda).value if col_azienda else None
        out.append((r, str(nome).strip() if nome else f"Riga {r}"))
    return out


def leggi_survey(source, riga=3):
    """Legge una risposta Qualtrics.

    Struttura attesa: riga 1 = codici domanda, riga 2 = testi, righe 3+ = risposte.
    """
    ws = _foglio(source, ["Sheet0"])
    colonne = {}
    for c in range(1, ws.max_column + 1):
        chiave = ws.cell(1, c).value
        if chiave:
            colonne[str(chiave).strip()] = c

    risposte, testi, col_origine = {}, {}, {}
    for chiave, c in colonne.items():
        if not re.fullmatch(r"(CQ|PQ|SQ|MQ)\d+", chiave):
            continue
        grezzo = ws.cell(riga, c).value
        try:
            risposte[chiave] = float(str(grezzo).replace(",", "."))
        except (TypeError, ValueError):
            risposte[chiave] = None
        testo = str(ws.cell(2, c).value or "")
        testi[chiave] = re.sub(r"\s+", " ", testo.replace("ⓘ", "")).strip()
        col_origine[chiave] = get_column_letter(c)

    if not risposte:
        raise ErroreInput(
            "nessuna colonna con codice domanda (CQ1, PQ1, SQ1, MQ1...) trovata "
            "nella riga 1. Attesa la struttura dell'export Qualtrics."
        )

    meta = {}
    for chiave, campo in META_QUALTRICS.items():
        c = colonne.get(chiave)
        meta[campo] = ws.cell(riga, c).value if c else None

    return {
        "azienda": str(meta["azienda"]).strip() if meta["azienda"] else f"Riga {riga}",
        "piva": meta["piva"],
        "dipendenti": meta["dipendenti"],
        "settore": meta["settore"],
        "dimensione": classifica_dimensione(meta["dipendenti"]),
        "riga": riga,
        "risposte": risposte,
        "testi": testi,
        "col_origine": col_origine,
    }


def calcola_x(survey, mappatura=MAPPATURA, codici=CODICI):
    """Score survey per tema.

        X(tema) = somma(risposta x peso x quota) / somma(peso x quota)

    Quota = 100% sul Cod. 1 se il Cod. 2 e' vuoto, altrimenti 70% / 30%.
    Sono escluse le domande MAT, quelle senza Cod. 1, senza peso o con risposta
    non valida. Restituisce (score_per_tema, dettaglio_riga_per_riga).
    """
    num, den, dettaglio = {}, {}, []
    for d in mappatura:
        peso = d["peso_pmi"] if survey["dimensione"] == "PMI" else d["peso_gr"]
        risposta = survey["risposte"].get(d["cod"])

        if d["standard"] == "MAT":
            motivo = "MAT - esclusa dal calcolo matrice"
        elif not d["cod1"]:
            motivo = "Cod. 1 assente"
        elif not isinstance(peso, (int, float)):
            motivo = "Peso assente/non numerico"
        elif not isinstance(risposta, float):
            motivo = "Risposta assente/non numerica"
        elif not 1 <= risposta <= 4:
            motivo = "Risposta fuori scala"
        else:
            motivo = "Inclusa"

        inclusa = motivo == "Inclusa"
        a1 = (1 if not d["cod2"] else 0.7) if inclusa else 0
        a2 = (0 if not d["cod2"] else 0.3) if inclusa else 0

        if inclusa:
            t1 = codici.get(d["cod1"], "Codice non mappato")
            num[t1] = num.get(t1, 0) + risposta * peso * a1
            den[t1] = den.get(t1, 0) + peso * a1
            if a2:
                t2 = codici.get(d["cod2"], "Codice non mappato")
                num[t2] = num.get(t2, 0) + risposta * peso * a2
                den[t2] = den.get(t2, 0) + peso * a2

        dettaglio.append({**d, "peso": peso, "risposta": risposta, "alloc1": a1,
                          "alloc2": a2, "inclusa": SI if inclusa else "No",
                          "motivo": motivo,
                          "tema1": codici.get(d["cod1"], ""),
                          "tema2": codici.get(d["cod2"], "")})

    score = {t: num[t] / den[t] for t in num if den.get(t)}
    return score, dettaglio


# --------------------------------------------------------- asse Y: reporting
def leggi_reporting(source):
    """Legge il foglio matrice_y: una riga per Azienda x Tema, score intero 1-4."""
    ws = _foglio(source, ["matrice_y"])
    intestazioni = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    attese = ["azienda", "tema reporting", "score reporting"]
    mancanti = [a for a in attese if a not in intestazioni]
    if mancanti:
        raise ErroreInput(
            f"colonne mancanti nel foglio dell'asse Y: {', '.join(mancanti)}. "
            "Attese: Azienda, Dimensione, Tema reporting, Score reporting, "
            "Usato matrice, Chiave lookup."
        )

    def cella(r, nome):
        c = intestazioni.get(norm(nome))
        return ws.cell(r, c).value if c else None

    righe = []
    for r in range(2, ws.max_row + 1):
        azienda = cella(r, "Azienda")
        tema = cella(r, "Tema reporting")
        if not azienda or not tema:
            continue
        score = cella(r, "Score reporting")
        try:
            score = float(str(score).replace(",", "."))
        except (TypeError, ValueError):
            score = None
        usato = cella(r, "Usato matrice")
        righe.append({
            "azienda": str(azienda).strip(),
            "dimensione": cella(r, "Dimensione"),
            "tema": str(tema).strip(),
            "score": score,
            "usato": SI if usato is None else usato,
            "chiave": cella(r, "Chiave lookup") or f"{azienda}|{tema}",
            "riga_fonte": cella(r, "Riga fonte dataset"),
        })
    if not righe:
        raise ErroreInput("nessuna riga di content analysis trovata per l'asse Y.")
    return righe


def aziende_reporting(reporting, solo_usate=True):
    """Aziende presenti nell'asse Y, in ordine di prima comparsa."""
    out = []
    for r in reporting:
        if solo_usate and norm(r["usato"]) != norm(SI):
            continue
        if r["azienda"] not in out:
            out.append(r["azienda"])
    return out


def score_y(reporting, azienda=None):
    """Mappa tema -> score reporting per l'azienda indicata."""
    return {r["tema"]: r["score"] for r in reporting
            if (azienda is None or r["azienda"] == azienda) and r["score"] is not None}


# ---------------------------------------------------------------- quadranti
def livello(v):
    """Scala di maturita' 1-4 secondo la metodologia."""
    if v is None or v == "":
        return ""
    if v <= 1.5:
        return "1 - Deriva - Discontinuo"
    if v <= 2.5:
        return "2 - Rotta - Strutturato"
    if v <= 3.5:
        return "3 - Orbita - Integrato"
    return "4 - Gravità - Sistemico / Consistente"


def quadrante(x, y, soglia=SOGLIA):
    alta_x, alta_y = x > soglia, y > soglia
    if alta_x and alta_y:
        return "Allineamento virtuoso"
    if alta_x and not alta_y:
        return "Potenziale nascosto"
    if not alta_x and alta_y:
        return "Sovraesposizione"
    return "Area fragile"


INTERPRETAZIONE = {
    "Allineamento virtuoso": "Area di coerenza: pratiche DEIA mature e disclosure strutturata risultano allineate.",
    "Potenziale nascosto": "Area di sottorendicontazione: pratiche più mature della disclosure esterna.",
    "Sovraesposizione": "Area di disallineamento comunicativo: la disclosure appare più avanzata della maturità praticata.",
    "Area fragile": "Area prioritaria di sviluppo: pratiche e disclosure risultano entrambe deboli.",
}

PRIORITA = {
    "Area fragile": "Alta - rafforzare congiuntamente pratiche e disclosure",
    "Sovraesposizione": "Alta - validare evidenze interne e rafforzare pratiche/policy",
    "Potenziale nascosto": "Media-Alta - migliorare qualità, granularità e completezza della disclosure",
    "Allineamento virtuoso": "Bassa - consolidare e usare come benchmark interno",
}


def costruisci_matrice(survey, reporting, azienda_y=None, soglia=SOGLIA):
    """Tabella finale: una riga per tema con X, Y, quadrante e commento.

    I temi senza dato su uno dei due assi sono restituiti con quadrante vuoto,
    cosi' da rendere visibile la lacuna invece di nasconderla.
    """
    x_map, dettaglio = calcola_x(survey)
    if azienda_y is None:
        disponibili = aziende_reporting(reporting) or aziende_reporting(reporting, False)
        azienda_y = disponibili[0] if disponibili else None
    y_map = score_y(reporting, azienda_y)

    righe = []
    for cod, tema in CODICI.items():
        x, y = x_map.get(tema), y_map.get(tema)
        q = quadrante(x, y, soglia) if x is not None and y is not None else ""
        righe.append({
            "Codice": cod,
            "Tema": tema,
            "X praticata": x,
            "Livello X": livello(x),
            "Y comunicata": y,
            "Livello Y": livello(y),
            "Differenza Y-X": (y - x) if x is not None and y is not None else None,
            "Maturità X": ("Alta" if x > soglia else "Bassa") if x is not None else "",
            "Maturità Y": ("Alta" if y > soglia else "Bassa") if y is not None else "",
            "Quadrante": q,
            "Interpretazione": INTERPRETAZIONE.get(q, "Dato mancante su almeno un asse"),
            "Priorità": PRIORITA.get(q, ""),
        })
    return righe, dettaglio, azienda_y
