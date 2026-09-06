# -*- coding: utf-8 -*-
"""
Generatore Excel della Matrice di trasparenza DEIA.

Produce un workbook con formule vive (ricalcolabili in Excel) a partire da:

    qualtrix_output.xlsx    -> asse X, maturita' DEIA praticata (survey)
    elaborato_giorgia.xlsx  -> asse Y, maturita' DEIA comunicata (reporting)

La metodologia e il calcolo stanno in deia_core.py, condivisi con app.py.

Uso:  python build_matrice.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from deia_core import (CODICI, MAPPATURA, SI, SOGLIA, calcola_x, leggi_reporting,
                       leggi_survey, livello, norm, quadrante)

BASE = Path(__file__).resolve().parent
F_SURVEY = BASE / "qualtrix_output.xlsx"
F_REPORT = BASE / "elaborato_giorgia.xlsx"
F_OUT = BASE / "Matrice_trasparenza_ricostruita.xlsx"

# ---------------------------------------------------------------- stile
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
T_FONT = Font(bold=True, size=13, color="1F3864")
B_FONT = Font(bold=True, size=10)
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
Q_FILL = {
    "Sovraesposizione": PatternFill("solid", fgColor="FCE4D6"),
    "Allineamento virtuoso": PatternFill("solid", fgColor="E2EFDA"),
    "Area fragile": PatternFill("solid", fgColor="F2F2F2"),
    "Potenziale nascosto": PatternFill("solid", fgColor="FFF2CC"),
}


# ------------------------------------------------- 5. scrittura workbook
def intesta(ws, riga, valori, larghezze=None):
    for i, v in enumerate(valori, start=1):
        c = ws.cell(riga, i, v)
        c.fill, c.font, c.border = H_FILL, H_FONT, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[riga].height = 30
    if larghezze:
        for i, w in enumerate(larghezze, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def scrivi(codici, domande, survey, reporting, dettaglio, score_x):
    temi = list(codici.values())
    wb = Workbook()

    # ---------- Note e assunzioni (tabella codici in A23:B32: NON spostare)
    ws = wb.active
    ws.title = "Note e assunzioni"
    ws["A1"], ws["A1"].font = "Note e assunzioni", T_FONT
    ws["A2"] = ("Parametri e regole della matrice. La tabella codice->tema in "
                "A23:B32 e' richiamata dalle formule del foglio Mappatura A-B-C.")
    note = [
        ("Fonte asse X (praticata)", f"{F_SURVEY.name}, foglio Sheet0, riga {survey['riga']}"),
        ("Fonte asse Y (comunicata)", f"{F_REPORT.name}, foglio matrice_y"),
        ("Fonte metodologia", "deia_core.py (mappatura, pesi, soglie)"),
        ("Caso survey", f"{survey['azienda']} - {survey['dipendenti']} dipendenti"),
        ("Dimensione / peso usato", f"{survey['dimensione']} -> colonna Peso "
                                    f"{'PMI' if survey['dimensione'] == 'PMI' else 'GRANDI'}"),
        ("Caso reporting", sorted({r['azienda'] for r in reporting if norm(r['usato']) == norm(SI)})),
        ("Regola MAT", "Le domande con Cod. standard = MAT sono escluse dalla matrice."),
        ("Regola allocazione", "Cod. 1 = 100% se Cod. 2 e' vuoto; Cod. 1 = 70% e Cod. 2 = 30% se Cod. 2 e' valorizzato."),
        ("Formula asse X", "Score = somma(risposta x peso x quota) / somma(peso x quota)."),
        ("Formula asse Y", "Valore intero 1-4 di content analysis per tema, non mediato (SUMIF su chiave Azienda|Tema)."),
        ("Soglia quadranti", f"Bassa maturita' <= {SOGLIA:.2f}; alta maturita' > {SOGLIA:.2f}."),
        ("Scala livello", "1: 1-1,50; 2: 1,51-2,50; 3: 2,51-3,50; 4: >3,50."),
        ("Esclusioni rilevate", "Vedi colonna 'Motivo esclusione / stato' in Mappatura A-B-C."),
    ]
    ws["A3"], ws["B3"] = "Elemento", "Assunzione / nota"
    for c in ("A3", "B3"):
        ws[c].fill, ws[c].font = H_FILL, H_FONT
    for i, (k, v) in enumerate(note, start=4):
        ws.cell(i, 1, k).font = B_FONT
        ws.cell(i, 2, ", ".join(v) if isinstance(v, list) else v)
    ws["A22"], ws["B22"] = "Codice tema", "Tema reporting"
    for c in ("A22", "B22"):
        ws[c].fill, ws[c].font = H_FILL, H_FONT
    for i, (cod, tema) in enumerate(codici.items(), start=23):
        ws.cell(i, 1, cod)
        ws.cell(i, 2, tema)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    for r in range(4, 33):
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    # ---------- Input survey A (asse X grezzo)
    ws = wb.create_sheet("Input survey A")
    ws["A1"], ws["A1"].font = "Input survey A - risposte Qualtrics (asse X)", T_FONT
    meta = [("File origine", F_SURVEY.name), ("Foglio origine", "Sheet0"),
            ("Riga origine", survey["riga"]), ("Ragione sociale", survey["azienda"]),
            ("Tipo impresa", survey["dimensione"]),
            ("Peso usato", "Peso PMI" if survey["dimensione"] == "PMI" else "Peso GRANDI")]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = B_FONT
        ws.cell(i, 2, v)
    # B7 = tipo impresa, B8 = peso usato: richiamati dalle formule
    intesta(ws, 11, ["Codice domanda", "Testo domanda", "Risposta survey",
                     "Col. Qualtrics", "In mappatura"], [16, 95, 14, 14, 13])
    codici_mappati = {d["cod"] for d in domande}
    riga = 12
    for cod in survey["risposte"]:
        ws.cell(riga, 1, cod)
        ws.cell(riga, 2, survey["testi"].get(cod, "")).alignment = Alignment(wrap_text=True)
        ws.cell(riga, 3, survey["risposte"][cod])
        ws.cell(riga, 4, survey["col_origine"].get(cod))
        ws.cell(riga, 5, SI if cod in codici_mappati else "No")
        riga += 1
    ultima_survey = riga - 1
    ws.freeze_panes = "A12"

    # ---------- Input reporting B (asse Y)
    ws = wb.create_sheet("Input reporting B")
    ws["A1"], ws["A1"].font = "Input reporting B - content analysis (asse Y)", T_FONT
    ws["A3"], ws["B3"] = "File origine", F_REPORT.name
    ws["A4"], ws["B4"] = "Foglio origine", "matrice_y"
    ws["A3"].font = ws["A4"].font = B_FONT
    intesta(ws, 8, ["Azienda", "Dimensione", "Tema reporting", "Score reporting",
                    "Usato matrice", "Chiave lookup", "Riga fonte dataset"],
            [26, 16, 26, 15, 14, 40, 18])
    riga = 9
    for r in reporting:
        ws.cell(riga, 1, r["azienda"])
        ws.cell(riga, 2, r["dimensione"])
        ws.cell(riga, 3, r["tema"])
        ws.cell(riga, 4, r["score"])
        ws.cell(riga, 5, r["usato"])
        ws.cell(riga, 6, r["chiave"])
        ws.cell(riga, 7, r["riga_fonte"])
        riga += 1
    ultima_rep = riga - 1
    ws.freeze_panes = "A9"

    # ---------- Mappatura A-B-C (motore di calcolo dell'asse X)
    ws = wb.create_sheet("Mappatura A-B-C")
    ws["A1"], ws["A1"].font = "Mappatura A-B-C", T_FONT
    ws["A2"] = ("Mappatura domanda survey -> tema reporting con pesi, quote 70/30 "
                "e contributi al numeratore/denominatore dell'asse X.")
    intesta(ws, 4, ["Col. Qualtrics", "N.", "Cod. domanda", "Framework diversity",
                    "Sottogruppo", "Cod. standard", "Cod. 1", "Cod. 2", "Tema 1",
                    "% alloc. 1", "Tema 2", "% alloc. 2", "Peso PMI", "Peso GRANDI",
                    "Peso usato", "Risposta survey", "Inclusa matrice",
                    "Motivo esclusione / stato", "Contributo num. tema 1",
                    "Contributo den. tema 1", "Contributo num. tema 2",
                    "Contributo den. tema 2"],
            [14, 6, 13, 15, 22, 13, 9, 9, 22, 11, 22, 11, 10, 12, 11, 13, 12, 30,
             12, 12, 12, 12])

    r0 = 5
    sr = f"'Input survey A'!$C$12:$C${ultima_survey}"
    sa = f"'Input survey A'!$A$12:$A${ultima_survey}"
    for i, d in enumerate(domande):
        r = r0 + i
        ws.cell(r, 1, survey["col_origine"].get(d["cod"], "n.d."))
        ws.cell(r, 2, i + 1)
        ws.cell(r, 3, d["cod"])
        ws.cell(r, 4, d["framework"])
        ws.cell(r, 5, d["sottogruppo"])
        ws.cell(r, 6, d["standard"])
        ws.cell(r, 7, d["cod1"])
        ws.cell(r, 8, d["cod2"])
        ws.cell(r, 9, f'=IF(G{r}="","",IFERROR(VLOOKUP(G{r},\'Note e assunzioni\'!$A$23:$B$32,2,FALSE),"Codice non mappato"))')
        ws.cell(r, 10, f'=IF(Q{r}="{SI}",IF(H{r}="",1,0.7),0)')
        ws.cell(r, 11, f'=IF(H{r}="","",IFERROR(VLOOKUP(H{r},\'Note e assunzioni\'!$A$23:$B$32,2,FALSE),"Codice non mappato"))')
        ws.cell(r, 12, f'=IF(Q{r}="{SI}",IF(H{r}="",0,0.3),0)')
        ws.cell(r, 13, d["peso_pmi"])
        ws.cell(r, 14, d["peso_gr"])
        ws.cell(r, 15, f'=IF(\'Input survey A\'!$B$7="PMI",IF(M{r}="","",M{r}),IF(N{r}="","",N{r}))')
        ws.cell(r, 16, f'=IFERROR(INDEX({sr},MATCH(C{r},{sa},0)),"")')
        ws.cell(r, 17, f'=IF(F{r}="MAT","No",IF(G{r}="","No",IF(O{r}="","No",'
                       f'IF(OR(NOT(ISNUMBER(P{r})),P{r}<1,P{r}>4),"No","{SI}"))))')
        ws.cell(r, 18, f'=IF(F{r}="MAT","MAT - esclusa dal calcolo matrice",'
                       f'IF(G{r}="","Cod. 1 assente",IF(O{r}="","Peso assente/non numerico",'
                       f'IF(NOT(ISNUMBER(P{r})),"Risposta assente/non numerica",'
                       f'IF(OR(P{r}<1,P{r}>4),"Risposta fuori scala","Inclusa")))))')
        ws.cell(r, 19, f'=IF(Q{r}="{SI}",P{r}*O{r}*J{r},0)')
        ws.cell(r, 20, f'=IF(Q{r}="{SI}",O{r}*J{r},0)')
        ws.cell(r, 21, f'=IF(Q{r}="{SI}",P{r}*O{r}*L{r},0)')
        ws.cell(r, 22, f'=IF(Q{r}="{SI}",O{r}*L{r},0)')
        for c in range(1, 23):
            ws.cell(r, c).border = BORDER
        for c in (19, 20, 21, 22):
            ws.cell(r, c).number_format = "0.000"
    r1 = r0 + len(domande) - 1
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:V{r1}"

    # ---------- Calcoli matrice C
    ws = wb.create_sheet("Calcoli matrice C")
    ws["A1"], ws["A1"].font = "Calcoli matrice C", T_FONT
    ws["A2"] = "Confronto tema per tema: asse X praticata (survey) vs asse Y comunicata (reporting)."
    aziende_y = sorted({r["azienda"] for r in reporting if norm(r["usato"]) == norm(SI)})
    ws["A3"], ws["B3"] = "Caso survey (X)", survey["azienda"]
    ws["A4"], ws["B4"] = "Caso reporting (Y)", aziende_y[0] if aziende_y else ""
    ws["A5"], ws["B5"] = "Soglia alta/bassa", SOGLIA
    ws["A6"], ws["B6"] = "Regola alta/bassa", f"Alta > {SOGLIA:.2f}; Bassa <= {SOGLIA:.2f}"
    for c in ("A3", "A4", "A5", "A6"):
        ws[c].font = B_FONT
    intesta(ws, 8, ["Tema reporting", "Codice", "Score survey X (praticata)",
                    "Livello survey", "Score reporting Y (comunicata)", "Livello reporting",
                    "Differenza Y-X", "Maturita' survey", "Maturita' reporting",
                    "Quadrante matrice", "Interpretazione sintetica", "Priorita' intervento",
                    "Denominatore survey", "Check range"],
            [24, 9, 15, 24, 15, 24, 12, 13, 13, 22, 62, 52, 14, 12])

    m = "'Mappatura A-B-C'"
    num1 = f"SUMIF({m}!$I${r0}:$I${r1},$A{{r}},{m}!$S${r0}:$S${r1})"
    num2 = f"SUMIF({m}!$K${r0}:$K${r1},$A{{r}},{m}!$U${r0}:$U${r1})"
    den1 = f"SUMIF({m}!$I${r0}:$I${r1},$A{{r}},{m}!$T${r0}:$T${r1})"
    den2 = f"SUMIF({m}!$K${r0}:$K${r1},$A{{r}},{m}!$V${r0}:$V${r1})"
    liv = ('=IF({c}{r}="","",IF({c}{r}<=1.5,"1 - Deriva - Discontinuo",'
           'IF({c}{r}<=2.5,"2 - Rotta - Strutturato",IF({c}{r}<=3.5,'
           '"3 - Orbita - Integrato","4 - Gravità - Sistemico / Consistente"))))')

    riga_tema = {}
    for i, tema in enumerate(temi):
        r = 9 + i
        riga_tema[tema] = r
        cod = [k for k, v in codici.items() if v == tema][0]
        ws.cell(r, 1, tema)
        ws.cell(r, 2, cod)
        ws.cell(r, 3, f'=IFERROR(({num1.format(r=r)}+{num2.format(r=r)})/'
                      f'({den1.format(r=r)}+{den2.format(r=r)}),"")')
        ws.cell(r, 4, liv.format(c="C", r=r))
        ws.cell(r, 5, f"=SUMIF('Input reporting B'!$F$9:$F${ultima_rep},$B$4&\"|\"&$A{r},"
                      f"'Input reporting B'!$D$9:$D${ultima_rep})")
        ws.cell(r, 6, liv.format(c="E", r=r))
        ws.cell(r, 7, f'=IF(OR(C{r}="",E{r}=""),"",E{r}-C{r})')
        ws.cell(r, 8, f'=IF(C{r}="","",IF(C{r}>$B$5,"Alta","Bassa"))')
        ws.cell(r, 9, f'=IF(E{r}="","",IF(E{r}>$B$5,"Alta","Bassa"))')
        ws.cell(r, 10, f'=IF(AND(H{r}="Alta",I{r}="Alta"),"Allineamento virtuoso",'
                       f'IF(AND(H{r}="Alta",I{r}="Bassa"),"Potenziale nascosto",'
                       f'IF(AND(H{r}="Bassa",I{r}="Alta"),"Sovraesposizione","Area fragile")))')
        ws.cell(r, 11, f'=IF(J{r}="Allineamento virtuoso","Area di coerenza: pratiche DEIA mature e disclosure strutturata risultano allineate.",'
                       f'IF(J{r}="Potenziale nascosto","Area di sottorendicontazione: pratiche più mature della disclosure esterna.",'
                       f'IF(J{r}="Sovraesposizione","Area di disallineamento comunicativo: la disclosure appare più avanzata della maturità praticata.",'
                       f'"Area prioritaria di sviluppo: pratiche e disclosure risultano entrambe deboli.")))')
        ws.cell(r, 12, f'=IF(J{r}="Area fragile","Alta - rafforzare congiuntamente pratiche e disclosure",'
                       f'IF(J{r}="Sovraesposizione","Alta - validare evidenze interne e rafforzare pratiche/policy",'
                       f'IF(J{r}="Potenziale nascosto","Media-Alta - migliorare qualità, granularità e completezza della disclosure",'
                       f'"Bassa - consolidare e usare come benchmark interno")))')
        ws.cell(r, 13, f'=({den1.format(r=r)}+{den2.format(r=r)})')
        ws.cell(r, 14, f'=IF(AND(C{r}>=1,C{r}<=4,E{r}>=1,E{r}<=4),"OK","ERRORE")')
        for c in range(1, 15):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")
        for c in (3, 5, 7, 13):
            ws.cell(r, c).number_format = "0.00"
    ultimo_calcolo = 8 + len(temi)
    ws.freeze_panes = "C9"

    # ---------- Matrice trasparenza
    ws = wb.create_sheet("Matrice trasparenza")
    ws["A1"], ws["A1"].font = "Matrice trasparenza", T_FONT
    ws["A2"] = "Posizionamento dei temi nei quattro quadranti."
    ws["A3"], ws["B3"] = "Asse X", "Maturita' DEIA praticata (survey Qualtrics)"
    ws["A4"], ws["B4"] = "Asse Y", "Maturita' DEIA comunicata (content analysis reporting)"
    ws["A5"], ws["B5"] = "Soglia alta/bassa", "='Calcoli matrice C'!$B$5"
    for c in ("A3", "A4", "A5"):
        ws[c].font = B_FONT

    ws["B8"] = f"Bassa maturita' survey (<={SOGLIA:.2f})"
    ws["C8"] = f"Alta maturita' survey (>{SOGLIA:.2f})"
    ws["A9"] = f"Alta maturita' reporting (>{SOGLIA:.2f})"
    ws["A10"] = f"Bassa maturita' reporting (<={SOGLIA:.2f})"
    for c in ("B8", "C8", "A9", "A10"):
        ws[c].fill, ws[c].font = H_FILL, H_FONT
        ws[c].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # colonne di appoggio P:S -> nome tema se appartiene al quadrante
    quadranti = {"B9": ("SOVRAESPOSIZIONE", "Sovraesposizione", 16),
                 "C9": ("ALLINEAMENTO VIRTUOSO", "Allineamento virtuoso", 17),
                 "B10": ("AREA FRAGILE", "Area fragile", 18),
                 "C10": ("POTENZIALE NASCOSTO", "Potenziale nascosto", 19)}
    for i in range(len(temi)):
        r = 16 + i
        rc = 9 + i
        for _, (_, nome, col) in quadranti.items():
            ws.cell(r, col, f'=IF(\'Calcoli matrice C\'!$J${rc}="{nome}",'
                            f'\'Calcoli matrice C\'!$A${rc},"")')
    ws.cell(15, 16, "app. Sovraesp.").font = B_FONT
    ws.cell(15, 17, "app. Allin.").font = B_FONT
    ws.cell(15, 18, "app. Fragile").font = B_FONT
    ws.cell(15, 19, "app. Potenz.").font = B_FONT
    for cella, (titolo, nome, col) in quadranti.items():
        L = get_column_letter(col)
        # _xlfn e' il prefisso richiesto da Excel per TEXTJOIN scritta da openpyxl
        rng = f"{L}16:{L}{15 + len(temi)}"
        ws[cella] = (f'="{titolo}"&CHAR(10)&IF(_xlfn.TEXTJOIN(CHAR(10),TRUE,{rng})="",'
                     f'"Nessun tema",_xlfn.TEXTJOIN(CHAR(10),TRUE,{rng}))')
        ws[cella].fill = Q_FILL[nome]
        ws[cella].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cella].border = BORDER
    for r in (9, 10):
        ws.row_dimensions[r].height = 130
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 42

    # tabella di sintesi + dati grafico (valori statici, ricalcolati dallo script)
    intesta(ws, 15, ["Tema", "Score survey X", "Score reporting Y", "Quadrante",
                     "Priorita'", "", "", "Tema chart", "X survey", "Y reporting"],
            [30, 15, 16, 22, 52, 3, 3, 26, 12, 12])
    score_y = {r["tema"]: r["score"] for r in reporting
               if norm(r["usato"]) == norm(SI)}
    sintesi = []
    for i, tema in enumerate(temi):
        r = 16 + i
        rc = riga_tema[tema]
        for col, src in ((1, "A"), (2, "C"), (3, "E"), (4, "J"), (5, "L")):
            ws.cell(r, col, f"='Calcoli matrice C'!{src}{rc}")
        x, y = score_x.get(tema), score_y.get(tema)
        ws.cell(r, 8, tema)
        ws.cell(r, 9, round(x, 6) if x is not None else None)
        ws.cell(r, 10, y)
        for c in (2, 3, 9, 10):
            ws.cell(r, c).number_format = "0.00"
        for c in list(range(1, 6)) + [8, 9, 10]:
            ws.cell(r, c).border = BORDER
        if x is not None and y is not None:
            q = quadrante(x, y)
            ws.cell(r, 4).fill = Q_FILL[q]
            sintesi.append((tema, x, y, q))
    ultima_sintesi = 15 + len(temi)

    # linee guida dei quadranti per il grafico
    base = ultima_sintesi + 3
    ws.cell(base, 8, "linea x=2,5")
    ws.cell(base, 9, SOGLIA)
    ws.cell(base, 10, 1)
    ws.cell(base + 1, 8, "linea x=2,5")
    ws.cell(base + 1, 9, SOGLIA)
    ws.cell(base + 1, 10, 4)
    ws.cell(base, 12, "linea y=2,5")
    ws.cell(base, 13, 1)
    ws.cell(base, 14, SOGLIA)
    ws.cell(base + 1, 12, "linea y=2,5")
    ws.cell(base + 1, 13, 4)
    ws.cell(base + 1, 14, SOGLIA)

    grafico = ScatterChart()
    grafico.title = "Matrice di trasparenza DEIA"
    grafico.style = 13
    grafico.height, grafico.width = 12, 17
    grafico.x_axis.title = "X - maturita' praticata (survey)"
    grafico.y_axis.title = "Y - maturita' comunicata (reporting)"
    grafico.x_axis.scaling.min = grafico.y_axis.scaling.min = 1
    grafico.x_axis.scaling.max = grafico.y_axis.scaling.max = 4
    punti = Series(Reference(ws, min_col=10, min_row=16, max_row=ultima_sintesi),
                   Reference(ws, min_col=9, min_row=16, max_row=ultima_sintesi),
                   title="Temi")
    punti.marker = Marker(symbol="circle", size=9)
    punti.graphicalProperties.line.noFill = True
    grafico.series.append(punti)
    for nome, cx, cy in (("soglia X", 9, 10), ("soglia Y", 13, 14)):
        s = Series(Reference(ws, min_col=cy, min_row=base, max_row=base + 1),
                   Reference(ws, min_col=cx, min_row=base, max_row=base + 1),
                   title=nome)
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line = LineProperties(solidFill="BFBFBF", w=12700,
                                                    prstDash="dash")
        grafico.series.append(s)
    ws.add_chart(grafico, f"A{ultima_sintesi + 5}")

    wb.save(F_OUT)
    return sintesi, ultimo_calcolo


def main():
    codici, domande = CODICI, MAPPATURA
    survey = leggi_survey(F_SURVEY)
    reporting = leggi_reporting(F_REPORT)
    score_x, dettaglio = calcola_x(survey)
    sintesi, _ = scrivi(codici, domande, survey, reporting, dettaglio, score_x)

    print(f"Caso X: {survey['azienda']} ({survey['dimensione']})")
    aziende_y = sorted({r['azienda'] for r in reporting if norm(r['usato']) == norm(SI)})
    print(f"Caso Y: {', '.join(aziende_y)}")
    escluse = [(d['cod'], d['motivo']) for d in dettaglio if d['inclusa'] != SI]
    print(f"Domande incluse: {len(dettaglio) - len(escluse)}/{len(dettaglio)} "
          f"- escluse: {escluse}")
    print(f"\n{'Tema':26s} {'X':>6s} {'Y':>6s}  Quadrante")
    for tema, x, y, q in sintesi:
        print(f"{tema:26s} {x:6.2f} {y:6.2f}  {q}")
    print(f"\nScritto: {F_OUT.name}")


if __name__ == "__main__":
    main()
