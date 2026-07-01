# -*- coding: utf-8 -*-
"""
Genera deia_mapping.xlsx a partire da deia_framework.json.

Fogli:
  - Istruzioni : come si usa
  - Mapping    : segnaposto | testo  (PRONTO: è ciò che l'app usa di default)
  - Assessment : id | tipo | nome | livello   (editabile, livelli 1-4)
  - Libreria   : id | tipo | nome | livello | testo  (tutti i 4 livelli)
  - Scala      : livello | nome | descrizione

Convenzione segnaposto nel PowerPoint:
  {{<id>.risultato}}  -> "<livello> - <NOME>"      (aree e sottogruppi)
  {{<id>.commento}}   -> commento dell'area al livello scelto
  {{<id>.attivita}}   -> attività del sottogruppo al livello scelto
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
fw = json.loads((BASE / "deia_framework.json").read_text(encoding="utf-8"))

scala = {str(s["livello"]): s["nome"] for s in fw["meta"]["scala"]}
ass = fw["assessment_esempio"]["aree"]


def risultato(lv):
    return f"{lv} - {scala[str(lv)]}"


# ---------------------------------------------------------------- raccolta dati
libreria = []      # id, tipo, nome, livello, testo
assessment = []    # id, tipo, nome, livello
mapping = []       # segnaposto, testo

for area in fw["aree"]:
    aid = area["id"]
    a_lv = ass[aid]["livello"]
    # libreria (commenti area, tutti i livelli)
    for lv in (1, 2, 3, 4):
        libreria.append((aid, "area", area["nome"], lv, area["commento"][str(lv)]))
    # assessment (livello scelto)
    assessment.append((aid, "area", area["nome"], a_lv))
    # mapping risolto
    mapping.append((f"{aid}.risultato", risultato(a_lv)))
    mapping.append((f"{aid}.commento", area["commento"][str(a_lv)]))

    for sg in area["sottogruppi"]:
        sid = sg["id"]
        s_lv = ass[aid]["sottogruppi"][sid]
        for lv in (1, 2, 3, 4):
            libreria.append((sid, "sottogruppo", sg["nome"], lv, sg["attivita"][str(lv)]))
        assessment.append((sid, "sottogruppo", sg["nome"], s_lv))
        mapping.append((f"{sid}.risultato", risultato(s_lv)))
        mapping.append((f"{sid}.attivita", sg["attivita"][str(s_lv)]))

# ------------------------------------------------------------------- stile xlsx
HDR_FILL = PatternFill("solid", fgColor="1F2A37")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER
    ws.freeze_panes = "A2"


def write_sheet(ws, headers, rows, widths, wrap_cols=()):
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws, len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            if cell.column in wrap_cols:
                cell.alignment = WRAP_TOP
            else:
                cell.alignment = Alignment(vertical="top")


wb = Workbook()

# --- Istruzioni
ws = wb.active
ws.title = "Istruzioni"
istruzioni = [
    ["ASSESSMENT DI MATURITÀ DEIA — File di mapping per l'app Streamlit"],
    [""],
    ["Come funziona"],
    ["1) Apri l'app Streamlit e carica il PowerPoint (template con segnaposto) e questo Excel."],
    ["2) L'app legge il foglio 'Mapping' e sostituisce ogni segnaposto {{...}} nel PPTX con il testo."],
    ["3) Scarichi il PowerPoint compilato."],
    [""],
    ["Segnaposto da usare nel PowerPoint (tra doppie graffe)"],
    ["{{<id>.risultato}}", "es. {{strategia.risultato}}  ->  '2 - ROTTA'"],
    ["{{<id>.commento}}", "solo AREE: il commento dell'area al livello scelto"],
    ["{{<id>.attivita}}", "solo SOTTOGRUPPI: le attività suggerite al livello scelto"],
    [""],
    ["Come cambiare i livelli (1-4)"],
    ["Opzione A — Veloce: modifica direttamente la colonna 'testo' nel foglio 'Mapping'."],
    ["Opzione B — Dinamica: cambia il numero di livello nel foglio 'Assessment' (colonna 'livello'),"],
    ["            poi nell'app spunta 'Rigenera mapping dai livelli'. L'app userà 'Assessment' + 'Libreria'."],
    [""],
    ["Foglio 'Libreria': contiene TUTTI e 4 i livelli per ogni area/sottogruppo (testi di riferimento)."],
    ["Foglio 'Scala': i nomi dei 4 livelli. Cambiali qui se la tua scala usa altri nomi."],
]
for r in istruzioni:
    ws.append(r)
ws["A1"].font = Font(bold=True, size=13, color="1F2A37")
for rk in (3, 8, 13):
    ws[f"A{rk}"].font = Font(bold=True, size=11, color="1F2A37")
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 70
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=False)

# --- Mapping (quello che l'app usa)
write_sheet(wb.create_sheet("Mapping"),
            ["segnaposto", "testo"], mapping,
            widths=[42, 110], wrap_cols=(2,))

# --- Assessment (editabile)
write_sheet(wb.create_sheet("Assessment"),
            ["id", "tipo", "nome", "livello"], assessment,
            widths=[40, 14, 42, 10])

# --- Libreria (riferimento completo)
write_sheet(wb.create_sheet("Libreria"),
            ["id", "tipo", "nome", "livello", "testo"], libreria,
            widths=[40, 14, 42, 9, 110], wrap_cols=(5,))

# --- Scala
write_sheet(wb.create_sheet("Scala"),
            ["livello", "nome", "descrizione"],
            [(s["livello"], s["nome"], s["descrizione"]) for s in fw["meta"]["scala"]],
            widths=[10, 18, 110], wrap_cols=(3,))

out = BASE / "deia_mapping.xlsx"
wb.save(out)
print(f"Mapping rows: {len(mapping)} | Libreria rows: {len(libreria)} | Assessment rows: {len(assessment)}")
print(f"Scritto: {out}")
